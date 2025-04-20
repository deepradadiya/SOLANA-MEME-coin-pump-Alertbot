import asyncio
import aiohttp
import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from typing import Dict
from telegram import Bot
import gc
import certifi
import ssl

ssl_context = ssl.create_default_context(cafile=certifi.where())

class SolanaTokenPriceFetcher:
    def __init__(self, telegram_token: str, telegram_chat_id: str):  # Fixed the constructor
        self.rpc_url = "https://api.mainnet-beta.solana.com"
        self.new_price_url = "https://fe-api.jup.ag/api/v1/prices"
        self.quote_address = "EPjFWdd5AufqSSqeM2qN1xzybapC8G4wEGGkZwyTDt1v"
        self.token_info_url = "https://fe-api.jup.ag/api/v1/tokens"
        self.file_name = "tokens.xlsx"
        self.telegram_token = telegram_token
        self.telegram_chat_id = telegram_chat_id
        self.bot = Bot(token=telegram_token)  # Ensure correct initialization

    def initialize_xlsx(self):
        """Create a new tokens.xlsx file with the appropriate columns if it doesn't exist."""
        if not os.path.exists(self.file_name):
            wb = Workbook()
            ws = wb.active
            ws.title = "Token Details"
            columns = ["Mint Address", "Balance", "Price USD", "Total Value"]
            for idx, col_name in enumerate(columns, start=1):
                ws[f"{get_column_letter(idx)}1"] = col_name
            wb.save(self.file_name)

    def update_xlsx(self, wallet_tokens: Dict[str, float], token_prices: Dict[str, float]):
        """Update the tokens.xlsx file with new token details and detect significant changes."""
        wb = load_workbook(self.file_name)
        ws = wb.active

        existing_data = {
            ws[f"A{row}"].value: {
                "row": row,
                "balance": ws[f"B{row}"].value,
                "total_value": ws[f"D{row}"].value,
            }
            for row in range(2, ws.max_row + 1)
        }

        significant_changes = []

        for mint, balance in wallet_tokens.items():
            if mint in token_prices:
                price_usd = token_prices[mint]
                total_value = balance * price_usd

                if mint in existing_data:
                    row = existing_data[mint]["row"]
                    if (
                        ws[f"B{row}"].value != balance
                        or ws[f"C{row}"].value != price_usd
                        or ws[f"D{row}"].value != total_value
                    ):
                        ws[f"B{row}"].value = balance
                        ws[f"C{row}"].value = price_usd
                        ws[f"D{row}"].value = total_value

                    if existing_data[mint]["total_value"]:
                        previous_value = existing_data[mint]["total_value"]
                        if total_value > 1.5 * previous_value or (
                            total_value > 21 and total_value > previous_value and total_value < 100) or total_value > 100:
                            significant_changes.append({
                                "mint": mint,
                                "new_total_value": total_value,
                            })
                else:
                    new_row = ws.max_row + 1
                    ws[f"A{new_row}"].value = mint
                    ws[f"B{new_row}"].value = balance
                    ws[f"C{new_row}"].value = price_usd
                    ws[f"D{new_row}"].value = total_value

        wb.save(self.file_name)
        return significant_changes

    async def get_wallet_tokens(self, wallet_address: str) -> Dict[str, float]:
        try:
            payload = {
                "jsonrpc": "2.0",
                "id": 1,
                "method": "getTokenAccountsByOwner",
                "params": [
                    wallet_address,
                    {"programId": "TokenkegQfeZyiNwAJbNbGKPFXCWuBvf9Ss623VQ5DA"},
                    {"encoding": "jsonParsed", "commitment": "confirmed"},
                ],
            }

            async with aiohttp.ClientSession() as session:
                async with session.post(self.rpc_url, json=payload, ssl=ssl_context) as response:
                    data = await response.json()

            token_balances = {}

            if "result" in data and "value" in data["result"]:
                for account in data["result"]["value"]:
                    parsed_info = account["account"]["data"]["parsed"]["info"]
                    mint_address = parsed_info["mint"]
                    token_amount = parsed_info["tokenAmount"]

                    balance = int(token_amount["amount"])
                    decimals = token_amount["decimals"]
                    readable_balance = balance / (10 ** decimals)

                    if readable_balance > 0:
                        token_balances[mint_address] = readable_balance

            return token_balances

        except Exception as e:
            print(f"Error fetching wallet tokens: {e}")
            return {}

    async def get_token_prices(self, session: aiohttp.ClientSession, mint_addresses: list) -> Dict[str, float]:
        try:
            list_address = ",".join(mint_addresses)
            url = f"{self.new_price_url}?list_address={list_address}"

            async with session.get(url) as response:
                data = await response.json()

            return data.get("prices", {})

        except Exception as e:
            print(f"Error fetching token prices: {e}")
            return {}

    async def fetch_all_token_prices(self, tokens: Dict[str, float]) -> Dict[str, float]:
        all_prices = {}

        async with aiohttp.ClientSession() as session:
            token_addresses = list(tokens.keys())
            for i in range(0, len(token_addresses), 2):
                batch_addresses = token_addresses[i:i + 2]
                prices = await self.get_token_prices(session, batch_addresses)
                all_prices.update(prices)

        return all_prices

    async def send_telegram_message(self, message: str, parse_mode: str = None):
        """Send a message to the Telegram bot."""
        await self.bot.send_message(chat_id=self.telegram_chat_id, text=message, parse_mode=parse_mode)


async def main_loop():
    wallet = "GyWkq2eg9DYZqNUiBGdfZghFbkx2ePL8N4zwnLVbAn27"
    telegram_token = "7774465414:AAGmtz-W6_dhdWLoTBA26Q_1wikyN8sFJ8M"
    telegram_chat_id = "7398132532"

    fetcher = SolanaTokenPriceFetcher(telegram_token, telegram_chat_id)  # Fixed instantiation

    fetcher.initialize_xlsx()
    current_iteration = 0
    while True:
        try:
            wallet_tokens = await fetcher.get_wallet_tokens(wallet)
            if wallet_tokens:
                token_prices = await fetcher.fetch_all_token_prices(wallet_tokens)
                significant_changes = fetcher.update_xlsx(wallet_tokens, token_prices)

                if significant_changes:
                    for token in significant_changes:
                        mint_address = token["mint"]
                        new_total_value = token["new_total_value"]

                        message = (
                            f"New Total Value: ${new_total_value:,.2f}\n"
                            f"Mint Address:\n<code> {mint_address}</code>\n"
                        )
                        await fetcher.send_telegram_message(message, parse_mode="HTML")
                        print(message)
            else:
                print("No tokens found in the wallet.")
            await asyncio.sleep(600)
        except Exception as e:
            print(f"Error during execution: {e}")
            await asyncio.sleep(600)
        finally:
            gc.collect()
            print(current_iteration)
            current_iteration += 1


if __name__ == "__main__":
    asyncio.run(main_loop())
